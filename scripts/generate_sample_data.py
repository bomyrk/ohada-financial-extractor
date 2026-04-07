"""
Generate synthetic OHADA financial statement Excel file for testing.

This creates a realistic sample Excel file matching the exact structure
of real DSF (Déclaration Statistique et Fiscale) files.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import numpy as np
from pathlib import Path


class OHADAExcelGenerator:
    """Generate sample OHADA Excel financial statement files."""

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def add_bilan_paysage(self):
        """Add Balance Sheet (Bilan Paysage) sheet."""
        ws = self.wb.create_sheet("BILAN PAYSAGE")
        
        # Headers
        ws['A1'] = "BILAN PAYSAGE"
        ws['A1'].font = Font(bold=True, size=14)
        
        periods = ["2023-12-31", "2024-12-31", "2025-12-31"]
        ws['B2'] = periods[0]
        ws['C2'] = periods[1]
        ws['D2'] = periods[2]
        
        # Asset accounts
        assets = [
            ("AD", "Immobilisations incorporelles", [(100000, 50000), (110000, 55000), (120000, 60000)]),
            ("AE", "Frais de développement", [(0, 0), (0, 0), (0, 0)]),
            ("AF", "Brevets et logiciels", [(50000, 25000), (55000, 27500), (60000, 30000)]),
            ("AZ", "Actif Immobilisé", [(500000, 150000), (550000, 165000), (600000, 180000)]),
            ("BA", "Actif circulant HAO", [(10000, 0), (12000, 0), (15000, 0)]),
            ("BB", "Stocks", [(200000, 0), (220000, 0), (240000, 0)]),
            ("BI", "Clients", [(300000, 0), (330000, 0), (360000, 0)]),
            ("BK", "Actif Circulant", [(1000000, 0), (1100000, 0), (1200000, 0)]),
            ("BS", "Banques", [(50000, 0), (60000, 0), (80000, 0)]),
            ("BT", "Trésorerie", [(100000, 0), (120000, 0), (150000, 0)]),
            ("BZ", "Total Actif", [(1600000, 150000), (1770000, 165000), (1950000, 180000)]),
        ]
        
        row = 3
        for code, label, periods_data in assets:
            ws[f'A{row}'] = code
            ws[f'B{row}'] = label
            for col_idx, (gross, amort) in enumerate(periods_data, 3):
                ws[f'{get_column_letter(col_idx+0)}{row}'] = gross
                ws[f'{get_column_letter(col_idx+1)}{row}'] = amort
                ws[f'{get_column_letter(col_idx+2)}{row}'] = gross - amort
            row += 1
        
        # Liability accounts
        liabilities = [
            ("CA", "Capital", [2000000, 2000000, 2000000]),
            ("CG", "Réserves", [100000, 120000, 140000]),
            ("CJ", "Résultat net", [150000, 160000, 170000]),
            ("CP", "Capitaux Propres", [2250000, 2280000, 2310000]),
            ("DA", "Emprunts", [500000, 400000, 300000]),
            ("DJ", "Fournisseurs", [600000, 700000, 800000]),
            ("DZ", "Total Passif", [1600000, 1770000, 1950000]),
        ]
        
        for code, label, values in liabilities:
            ws[f'A{row}'] = code
            ws[f'B{row}'] = label
            for col_idx, val in enumerate(values, 3):
                ws[f'{get_column_letter(col_idx)}{row}'] = val
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        for col in ['C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
        
        return ws

    def add_compte_resultat(self):
        """Add Income Statement (Compte de Résultat) sheet."""
        ws = self.wb.create_sheet("COMPTE DE RESULTAT")
        
        ws['A1'] = "COMPTE DE RESULTAT"
        ws['A1'].font = Font(bold=True, size=14)
        
        periods = ["2023-12-31", "2024-12-31", "2025-12-31"]
        for idx, period in enumerate(periods):
            ws[f'{get_column_letter(idx+2)}1'] = period
        
        income_items = [
            ("TA", "Ventes marchandises", [4000000, 4400000, 4840000]),
            ("XB", "Chiffre d'affaires", [5000000, 5500000, 6050000]),
            ("RC", "Achats matières", [2000000, 2200000, 2420000]),
            ("RK", "Charges personnel", [1000000, 1050000, 1100000]),
            ("XC", "Valeur ajoutée", [2000000, 2250000, 2530000]),
            ("XD", "Excédent brut exploitation", [1000000, 1200000, 1430000]),
            ("RL", "Dotations amortissements", [100000, 105000, 110000]),
            ("XE", "Résultat exploitation", [900000, 1095000, 1320000]),
            ("RM", "Charges financières", [50000, 55000, 60000]),
            ("XF", "Résultat financier", [50000, 55000, 60000]),
            ("XG", "Résultat ordinaire", [850000, 1040000, 1260000]),
            ("RS", "Impôts", [700000, 880000, 1090000]),
            ("XI", "Résultat Net", [150000, 160000, 170000]),
        ]
        
        row = 3
        for code, label, values in income_items:
            ws[f'A{row}'] = code
            ws[f'B{row}'] = label
            for col_idx, val in enumerate(values, 3):
                ws[f'{get_column_letter(col_idx)}{row}'] = val
            row += 1
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        for col in ['C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
        
        return ws

    def add_tableau_flux(self):
        """Add Cash Flow Statement (Tableau des Flux de Trésorerie) sheet."""
        ws = self.wb.create_sheet("TABLEAU DES FLUX DE TRESORERIE")
        
        ws['A1'] = "TABLEAU DES FLUX DE TRESORERIE"
        ws['A1'].font = Font(bold=True, size=14)
        
        periods = ["2023-12-31", "2024-12-31", "2025-12-31"]
        for idx, period in enumerate(periods):
            ws[f'{get_column_letter(idx+2)}1'] = period
        
        cashflow_items = [
            ("ZA", "Trésorerie initiale", [50000, 60000, 80000]),
            ("FA", "Capacité autofinancement", [300000, 350000, 400000]),
            ("FB", "Variation stocks", [-20000, -10000, -5000]),
            ("FC", "Variation créances", [-50000, -30000, -20000]),
            ("ZB", "Flux opérationnel", [230000, 310000, 375000]),
            ("FF", "Acquisitions immobilisé", [-100000, -110000, -120000]),
            ("FI", "Cessions immobilisé", [20000, 25000, 30000]),
            ("ZC", "Flux investissement", [-80000, -85000, -90000]),
            ("FJ", "Augmentation capital", [0, 0, 0]),
            ("FM", "Dividendes versés", [-100000, -110000, -120000]),
            ("ZD", "Flux capitaux propres", [-100000, -110000, -120000]),
            ("ZE", "Flux financement", [-50000, -35000, -20000]),
            ("ZH", "Trésorerie finale", [60000, 80000, 125000]),
        ]
        
        row = 3
        for code, label, values in cashflow_items:
            ws[f'A{row}'] = code
            ws[f'B{row}'] = label
            for col_idx, val in enumerate(values, 3):
                ws[f'{get_column_letter(col_idx)}{row}'] = val
            row += 1
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        for col in ['C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
        
        return ws

    def generate(self, output_path: str = None):
        """Generate the complete workbook."""
        if output_path is None:
            output_path = Path(__file__).parent.parent / "examples" / "data" / "sample_ohada_statement_2024.xlsx"
        
        # Create directory if needed
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        # Add all sheets
        self.add_bilan_paysage()
        self.add_compte_resultat()
        self.add_tableau_flux()
        
        # Save
        self.wb.save(output_path)
        print(f"✓ Generated sample file: {output_path}")
        return output_path


if __name__ == '__main__':
    generator = OHADAExcelGenerator()
    generator.generate()
    print("\nSample data generated successfully!")