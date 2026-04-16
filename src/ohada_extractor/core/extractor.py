"""
Main OHADA Financial Data Extractor

Handles Excel file parsing and financial data extraction according to
OHADA accounting standards.
"""

import logging
import openpyxl
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Union
from dateutil.relativedelta import relativedelta
from openpyxl.worksheet.worksheet import Worksheet

from .statement import FinancialStatement
from .schemas import OHADA_STATEMENTS, OHADA_NOTES, OTHER_ACCOUNTS


logger = logging.getLogger(__name__)


class FinancialExtractor:
    """
    Extract financial data from OHADA-compliant Excel financial statements.
    
    Supports:
    - Balance sheets (Bil-an Paysage) with Gross/Amortization/Net tracking
    - Income statements (Compte de Résultat)
    - Cash flow statements (Tableau des Flux de Trésorerie)
    - Multi-year consolidation
    - JSON export
    """

    # ---------------------------------------------------------
    # 1. INITIALIZATION
    # ---------------------------------------------------------
    def __init__(self):
        self.workbook = None
        self.raw_data = {}

    # ---------------------------------------------------------
    # 2. PUBLIC API (Top-level methods)
    # ---------------------------------------------------------
    def extract_from_excel(self, file_path: Union[str, Path]) -> 'FinancialStatement':
        """
        Extract financial data from a single OHADA Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            FinancialStatement object with extracted data
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If required sheets are missing
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self._validate_sheets()
        self._extract_all_statements()
        self._extract_all_notes()
        
        # Extract periods from headers
        periods = self._extract_periods()
        
        from .statement import FinancialStatement
        return FinancialStatement(
            asset_data=self.raw_data.get('assets'),
            liability_data=self.raw_data.get('liabilities'),
            income_data=self.raw_data.get('income'),
            cashflow_data=self.raw_data.get('cashflow'),
            other_data=self.raw_data.get('other'),
            notes=self.raw_data.get('notes'),
            periods=periods,
            file_path=str(file_path),
        )

    def extract_over_period(self, file_list: List[Union[str, Path]]) -> 'FinancialStatement':
        """
        Extract and consolidate financial data from multiple files across years.
        
        Args:
            file_list: List of Excel file paths
            
        Returns:
            Consolidated FinancialStatement
        """
        if len(file_list) < 2:
            raise ValueError("Minimum 2 files required for period analysis")

        combined_asset, combined_liabilities, combined_income, combined_cashflow, combined_other = (None,) * 5

        statements = []

        # Extract each file individually
        for file_path in file_list:
            statements.append(self.extract_from_excel(file_path))

        # Initialize combined arrays using the first statement
        first = statements[0]
        combined_asset = np.hstack((np.insert(first.asset_data.copy()[:, [0, -1]], [1],
                                              [np.nan, np.nan], axis=1), first.asset_data.copy()[:, 1:-1]))
        combined_liabilities = np.hstack((first.liability_data.copy()[:, [0, -1]], first.liability_data.copy()[:, 1:-1]))
        combined_income = np.hstack((first.income_data.copy()[:, [0, -1]], first.income_data.copy()[:, 1:-1]))
        combined_cashflow = np.hstack((first.cashflow_data.copy()[:, [0, -1]], first.cashflow_data.copy()[:, 1:-1]))
        combined_other = np.hstack((first.other_data.copy()[:, [0, -1]], first.other_data.copy()[:, 1:-1]))
        combined_notes = first.notes

        # Merge subsequent years
        for stmt in statements[1:]:
            combined_asset = self._merge_period_data(
                combined_asset, stmt.asset_data,
                combined_col=combined_asset.shape[1] - 1,
                new_col=4,
                label="asset"
            )

            combined_liabilities = self._merge_period_data(
                combined_liabilities, stmt.liability_data,
                combined_col=combined_liabilities.shape[1] - 1,
                new_col=2,
                label="liabilities"
            )

            combined_income = self._merge_period_data(
                combined_income, stmt.income_data,
                combined_col=combined_income.shape[1] - 1,
                new_col=2,
                label="income"
            )

            combined_cashflow = self._merge_period_data(
                combined_cashflow, stmt.cashflow_data,
                combined_col=combined_cashflow.shape[1] - 1,
                new_col=2,
                label="cashflow"
            )

            combined_other = self._merge_period_data(
                combined_other, stmt.other_data,
                combined_col=combined_other.shape[1] - 1,
                new_col=2,
                label="other"
            )

            combined_notes = self._merge_notes(
                combined_notes, stmt.notes
            )

        # Build periods list (sorted by year)
        periods = [stmt.periods[0] for stmt in statements]
        periods.insert(0, first.periods[1])

        # Replace None value with 0
        for tmp_data in [combined_asset, combined_liabilities, combined_income, combined_cashflow, combined_other]:
            tmp_data[np.where(tmp_data == None)] = 0

        # Return a proper FinancialStatement
        return FinancialStatement(
            asset_data=combined_asset,
            liability_data=combined_liabilities,
            income_data=combined_income,
            cashflow_data=combined_cashflow,
            other_data=combined_other,
            notes=combined_notes,
            periods=periods,
            file_path=";".join(str(p) for p in file_list)
        )

    # ---------------------------------------------------------
    # 3. CORE EXTRACTION PIPELINE
    # ---------------------------------------------------------
    def _validate_sheets(self):
        """Check that required sheets exist in workbook."""
        workbook_sheets = {' '.join(s.split()).lower() for s in self.workbook.sheetnames}
        required_sheets = {
            stmt.sheet_name.lower() 
            for stmt in OHADA_STATEMENTS.values()
        }
        missing = required_sheets - workbook_sheets
        if missing:
            raise ValueError(f"Missing required sheets: {missing}")
        logger.info(f"Workbook validated: all required sheets present")

    def _extract_all_statements(self):
        """Extract data from all financial statement sheets."""
        for key, stmt_config in OHADA_STATEMENTS.items():
            sheet = self._get_sheet(stmt_config.sheet_name)
            if sheet:
                data = self._extract_sheet_data(
                    sheet, 
                    stmt_config.start_account,
                    stmt_config.end_account,
                    stmt_config.account_count,
                    stmt_config.columns_idx,
                    stmt_config.lines_idx
                )
                if key == 'other_last_years':
                    data[:,0] = [item[1] for item in OTHER_ACCOUNTS]

                self.raw_data[key.split('_')[0]] = data

    def _get_sheet(self, sheet_name: str) -> Optional[Worksheet]:
        """Get worksheet by normalized name."""
        for ws in self.workbook.worksheets:
            if ' '.join(ws.title.split()).lower() == sheet_name.lower():
                return ws
        return None

    def _extract_sheet_data(
        self,
        sheet: Worksheet,
        start_code: str, 
        end_code: str, 
        length: int,
        columns: tuple[int, ...],
        lines: tuple[int, ...],
        strict: bool = True, pre_process: bool = False
    ) -> np.ndarray|tuple[np.ndarray, np.ndarray]:
        """
        Extract financial data from sheet between start and end codes.
        
        Args:
            sheet: Worksheet to extract from
            start_code: Starting account code (e.g., 'AD')
            end_code: Ending account code (e.g., 'BZ')
            length: Expected number of rows
            columns (tuple[int, ...], optional): the column indices to extract. if not provides, all columns will be
            extracts. Defaults to None.
            lines (tuple[int, ...], optional): The row indices to extract. if not provides, all rows will be extracts.
                        Defaults to None.
            strict (bool): If True, the function will make sure that the table extracted has exactly the same number of
                        rows as the number of rows specified in parameter length
            pre_process (bool): If True, the function will pre-process the data and return both raw data
                        and pre-processed data.
            
        Returns:
            NumPy array with extracted data
        """
        raw_data = []
        start = False
        liability_column_idx = 7

        for row in sheet.iter_rows(values_only=True):
            if row[0] is None and end_code not in [str(val).upper() for val in row[1:]]:
                continue
                
            # Check if we've found the start code
            if start_code.upper() in str(row).strip().upper():
                if start_code == 'CA' and row[liability_column_idx] is not None:
                    if row[liability_column_idx].strip().upper() == start_code.upper() :
                        start = True
                    elif start:
                        pass
                    else:
                        continue
                    raw_data.append(row[liability_column_idx:])
                else:
                    if row[0].strip().upper() == start_code.upper() :
                        start = True
                    elif start:
                        pass
                    else:
                        continue
                    raw_data.append(row)
            # Check if we've found the end code
            elif str(end_code).upper() in [str(val).strip().upper() for val in row]:
                if start_code == 'CA':
                    if start and str(end_code).upper() == str(row[liability_column_idx]).strip().upper():
                        raw_data.append(row[liability_column_idx:])
                        break
                else:
                    if start and (str(end_code).upper() == str(row[0]).strip().upper() or \
                            str(end_code).upper() == str(row[1:]).strip().upper()):
                        raw_data.append(row)
                        break
            # Extract data between start and end
            elif start:
                if start_code == 'CA':
                    if row[liability_column_idx] is None:
                        continue
                    else:
                        raw_data.append(row[liability_column_idx:])
                else:
                    raw_data.append(row)

        if (strict and len(raw_data) == length) or (strict is False and len(raw_data) <= length):
            if columns is None:
                if lines is None:
                    return np.array(raw_data)
                else:
                    if not pre_process:
                        return np.array(raw_data)[lines, :]
                    else:
                        return np.array(raw_data), np.array(raw_data)[:, lines]
            else:
                if lines is None:
                    if not pre_process:
                        return np.array(raw_data)[:, columns]
                    else:
                        return np.array(raw_data), np.array(raw_data)[:, columns]
                else:
                    if not pre_process:
                        return np.array(raw_data)[np.ix_(lines, columns)]
                    else:
                        return np.array(raw_data), np.array(raw_data)[np.ix_(lines, columns)]
        else:
            logger.warning(
                f"Expected {length} rows, got {len(raw_data)} "
                f"for range {start_code}-{end_code}"
            )
            return np.array(raw_data)

    def _extract_periods(self) -> List[str]:
        """Extract period dates from first available sheet."""
        for ws in self.workbook.worksheets:
            if ws.title.strip().lower() == "fiche r1":
                # Look for dates in header row (usually row 13 Fiche R1)
                cell = ws.cell(row=13, column=6)
                # Check if the cell has a value
                if cell.value:
                    # If the value is a date, format it and return
                    if isinstance(cell.value, datetime):
                        return [cell.value.strftime('%Y-%m-%d'),
                                (cell.value - relativedelta(years=1)).strftime('%Y-%m-%d')]  # Return as a string in YYYY-MM-DD format

                    # If the value is a string and can be converted to a date
                    try:
                        parsed_date = datetime.strptime(cell.value, '%Y-%m-%d')  # Adjust format if needed
                        return [
                            str(cell.value),  # Current string value
                            (parsed_date - relativedelta(years=1)).strftime('%Y-%m-%d')  # Date one year before
                        ]
                    except ValueError:
                        # Handle case where string can't be parsed as date
                        continue
        # Calculate the last day of the previous year
        last_day_last_year = datetime(datetime.now().year - 1, 12, 31)
        return [last_day_last_year.strftime('%Y-%m-%d'),
                (last_day_last_year - relativedelta(years=1)).strftime('%Y-%m-%d') ]  # Return formatted last day of last year

    # ---------------------------------------------------------
    # 4. NOTES EXTRACTION SUBSYSTEM
    # ---------------------------------------------------------
    def _extract_all_notes(self):
        """
        Extract all OHADA notes (annexes) defined in OHADA_NOTES.
        Handles:
        - optional sheets (skips missing ones)
        - special sheets (Fiche R2, NOTE 5, NOTE 30)
        - raw + preprocessed extraction
        - multi-key notes (e.g., Fiche R2)
        """
        self.raw_data['notes'] = {}

        # Normalize sheet names for lookup
        sheet_map = {
            ws.title.strip().lower(): ws
            for ws in self.workbook.worksheets
        }

        for note_id, cfg in OHADA_NOTES.items():
            sheet_key = cfg.sheet_name.lower().strip()

            # Skip missing sheets (notes are optional)
            if sheet_key not in sheet_map:
                logger.warning(f"Note sheet '{cfg.sheet_name}' not found — skipping.")
                continue

            sheet = sheet_map[sheet_key]

            # ---------------------------------------------------------
            # SPECIAL NOTES (custom logic)
            # ---------------------------------------------------------
            if cfg.is_special:

                # Extract raw data only
                raw_value = self._extract_sheet_data(
                    sheet,
                    cfg.start_marker,
                    cfg.end_marker,
                    cfg.expected_rows,
                    cfg.columns_idx,
                    cfg.lines_idx,
                    strict=False,
                    pre_process=False
                )

                # FICHE R2 special processing
                if cfg.sheet_name == 'Fiche R2':
                    preprocess_value = tuple(self._process_fiche_r2(raw_value))

                # NOTE 5 / NOTE 30 special preprocessing
                elif cfg.preprocess_rules:
                    preprocess_value = self._preprocess_note_data(
                        np.squeeze(raw_value),
                        columns=cfg.preprocess_rules.get('columns'),
                        refs=cfg.preprocess_rules.get('refs')
                    )

                else:
                    preprocess_value = None

            # ---------------------------------------------------------
            # STANDARD NOTES (raw + preprocessed)
            # ---------------------------------------------------------
            else:
                raw_value, preprocess_value = self._extract_sheet_data(
                    sheet,
                    cfg.start_marker,
                    cfg.end_marker,
                    cfg.expected_rows,
                    cfg.columns_idx,
                    cfg.lines_idx,
                    strict=True,
                    pre_process=True
                )

            # ---------------------------------------------------------
            # STORE RESULTS
            # ---------------------------------------------------------
            if isinstance(cfg.keys, str):
                # Single note
                self.raw_data['notes'][cfg.keys] = {
                    'name': cfg.names,
                    'raw_value': raw_value,
                    'preprocess_value': preprocess_value
                }

            else:
                # Multi-key note (e.g., Fiche R2)
                for i, key in enumerate(cfg.keys):
                    self.raw_data['notes'][key] = {
                        'name': cfg.names[i],
                        'raw_value': raw_value,
                        'preprocess_value': preprocess_value[i]
                    }

    # ---------------------------------------------------------
    # 5. NOTES HELPER METHODS
    # ---------------------------------------------------------
    def _preprocess_note_data(self, data, rows=None, columns=None, refs=None):
        """
        Preprocess note data by extracting specific rows/columns or filtering by reference labels.
        Converts None to 0 and returns a numeric numpy array.
        """
        # Replace None with 0
        data = np.where(data == None, 0, data)

        # Case 1: reference-based extraction
        if refs is not None:
            ref_indices = np.isin(data[:, 0], refs).nonzero()[0]
            if len(ref_indices) > 0:
                extracted = data[ref_indices]
                if columns is not None:
                    extracted = extracted[:, columns]
                else:
                    extracted = extracted[:, 1:]
                return extracted.astype(float)
            return np.array([])

        # Case 2: row slicing
        if rows is not None:
            data = data[rows]

        # Case 3: column slicing
        if columns is not None:
            data = data[:, columns]
        else:
            data = data[:, 1:]

        return data

    def _process_fiche_r2(self, raw_data):
        """
        Process Fiche R2 data into structured blocks:
        - Identification
        - Revenue breakdown
        """
        # Replace None with 0
        data = np.array([tuple(0 if x is None else x for x in row) for row in raw_data])

        # Block A: legal form, HQ, number of establishments
        a = self._convert_array(
            np.insert(data[np.ix_((0, 2, 3), (0, 5, 6, 7))], 1, 0, axis=1)
        )

        # Block B: tax regime, establishments abroad
        b = self._convert_array(
            np.insert(data[np.ix_((1, 4), (0, 5, 6))], (1, 2), 0, axis=1)
        )

        # Block C: first year of operation
        c = self._convert_array(data[np.ix_([5], (0, 5, 6, 7, 8, 9))])

        # Block D: ownership structure
        d = data[np.ix_((0, 1, 2), (11, 15))]

        stacked = np.vstack((a, b, c, d))

        # Revenue breakdown
        e = data[7:, (0, 5, 6, 7, 8, 9, 10, 14, 15)]
        income_breakdown = np.hstack((self._convert_array(e[:, :-2]), e[:, -2:]))

        return stacked, income_breakdown

    # ---------------------------------------------------------
    # 6. MULTI-YEAR CONSOLIDATION
    # ---------------------------------------------------------
    @staticmethod
    def _check_data_coherence(finStatY: np.ndarray, finStatLY: np.ndarray, i: int, j: int) -> bool:
        """
        Checks whether the financial data for a given column is consistent between the current year and the previous year.

        Args:
            finStatY (np.ndarray): Financial statements for the current year.
            finStatLY (np.ndarray): Financial statements for the previous year.
            i (int): Column index in `finStatY` to validate.
            j (int): Column index in `finStatLY` to validate.

        Returns:
            bool: True if the values in column `i` of the current year are identical
            to the values in column `j` of the previous year.
            False otherwise.
        """
        finStatY[finStatY[:, i] == None, i] = 0
        finStatLY[finStatLY[:, j] == None, j] = 0

        if np.array_equal(finStatY[:, i], finStatLY[:, j]):
            return True
        else:
            return False

    @staticmethod
    def _merge_period_data(
        combined: np.ndarray,
        new_data: np.ndarray,
        combined_col: int,
        new_col: int,
        label: str
    ) -> np.ndarray:
        """
        Merge multi-year financial data with coherence checking.

        Args:
            combined: Existing consolidated data
            new_data: Newly extracted data for the next year
            combined_col: Column index in `combined` to compare
            new_col: Column index in `new_data` to compare
            label: Name of the statement (for logging)

        Returns:
            Updated combined array (or unchanged if incoherent)
        """
        if FinancialExtractor._check_data_coherence(combined, new_data, combined_col, new_col):
            return np.hstack((combined, new_data[:, 1:-1]))
        else:
            logger.warning(f"Incoherence found in {label} data between years.")
            return combined

    @staticmethod
    def _merge_notes(notes1: dict, notes2: dict) -> dict:
        """
        Merge notes across years.
        - Keeps notes even if missing in one year
        - Handles raw_value and preprocess_value
        - Avoids duplicating the first column of raw_value
        - Handles tuple preprocess values (e.g., Fiche R2)
        """
        result = {}

        all_keys = set(notes1.keys()) | set(notes2.keys())

        for key in all_keys:
            n1 = notes1.get(key)
            n2 = notes2.get(key)

            # Case 1: note exists only in year 1
            if n1 is not None and n2 is None:
                result[key] = n1
                continue

            # Case 2: note exists only in year 2
            if n1 is None and n2 is not None:
                result[key] = n2
                continue

            # Case 3: note exists in both years → merge
            raw1 = n1["raw_value"]
            raw2 = n2["raw_value"]

            # Handle tuple preprocess values (Fiche R2)
            pre1 = n1["preprocess_value"]
            pre2 = n2["preprocess_value"]

            # Merge raw_value (avoid duplicating first column)
            merged_raw = np.hstack((raw1, raw2[:, 1:]))

            # Merge preprocess_value
            if isinstance(pre1, tuple):
                # Fiche R2 case: merge each element separately
                merged_pre = tuple(
                    np.hstack((pre1[i], pre2[i][:, 1:])) if isinstance(pre1[i], np.ndarray) else pre1[i]
                    for i in range(len(pre1))
                )
            else:
                merged_pre = np.hstack((pre1, pre2))

            result[key] = {
                "name": n1["name"],
                "raw_value": merged_raw,
                "preprocess_value": merged_pre
            }

        return result

    # ---------------------------------------------------------
    # 7. LOW-LEVEL UTILITIES
    # ---------------------------------------------------------
    def _is_nested_tuple(self, data)->bool:
        if isinstance(data, tuple):
            for element in data:
                if isinstance(element, tuple) or self._is_nested_tuple(element):
                    return True
        return False

    def _convert_array(self, data: np.ndarray) -> np.ndarray:
        """
        Convert a numpy array containing strings/integers into a 2-column array:
        - Column 0: original first column
        - Column 1: concatenation of remaining columns (as strings)
        """
        result = np.empty((data.shape[0], 2), dtype='object')

        # First column unchanged
        result[:, 0] = data[:, 0]

        # Combine remaining columns into a single string
        for i in range(data.shape[0]):
            result[i, 1] = ''.join(map(str, data[i, 1:]))

        return result







